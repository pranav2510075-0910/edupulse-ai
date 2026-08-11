from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

# ==========================================================
# CONFIGURATION
# ==========================================================

DATA_FOLDER = "data"
FILE_NAME = os.path.join(DATA_FOLDER, "feedback.xlsx")

# Create data folder automatically
os.makedirs(DATA_FOLDER, exist_ok=True)


# ==========================================================
# HOME PAGE
# ==========================================================

@app.route("/")
def home():
    return render_template("index.html")


# ==========================================================
# SAVE FEEDBACK
# ==========================================================

@app.route("/save", methods=["POST"])
def save():

    try:

        # --------------------------------------------------
        # Receive JSON from JavaScript
        # --------------------------------------------------

        data = request.get_json()

        if not data:

            return jsonify({
                "status": "error",
                "message": "No feedback data received."
            }), 400


        print("\n====================================")
        print("📥 NEW FEEDBACK RECEIVED")
        print("====================================")

        print(data)


        # --------------------------------------------------
        # Create Excel file if it doesn't exist
        # --------------------------------------------------

        if os.path.exists(FILE_NAME):

            wb = load_workbook(FILE_NAME)

            ws = wb.active

        else:

            wb = Workbook()

            ws = wb.active

            ws.title = "Student Feedback"


            # --------------------------------------------------
            # Excel Headers
            # --------------------------------------------------

            headers = [

                "Date & Time",

                # Student Details
                "Name",
                "Email",
                "Department",
                "Year",
                "Subject",
                "Faculty",

                # Quick Questions
                "Mood",
                "Class Rating",
                "Faculty Interaction",
                "Concept Clarity",
                "Doubt Encouragement",
                "Teaching Pace",

                # Deep Feedback
                "Favorite Part",
                "Difficult Topic",
                "Faculty Strength",
                "What Should Improve",
                "Preferred Teaching Method",
                "Suggestion",
                "Final Feedback"

            ]

            ws.append(headers)


        # --------------------------------------------------
        # Add new feedback row
        # --------------------------------------------------

        row = [

            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

            # Student Details
            data.get("name", ""),
            data.get("email", ""),
            data.get("department", ""),
            data.get("year", ""),
            data.get("subject", ""),
            data.get("faculty", ""),

            # Quick Questions
            data.get("mood", ""),
            data.get("classRating", ""),
            data.get("interaction", ""),
            data.get("clarity", ""),
            data.get("doubts", ""),
            data.get("pace", ""),

            # Deep Feedback
            data.get("favoritePart", ""),
            data.get("difficultTopic", ""),
            data.get("facultyStrength", ""),
            data.get("improvement", ""),
            data.get("teachingMethod", ""),
            data.get("suggestion", ""),
            data.get("finalFeedback", "")

        ]


        ws.append(row)


        # --------------------------------------------------
        # Make columns readable
        # --------------------------------------------------

        for column in ws.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:

                    cell_length = len(
                        str(cell.value)
                    )

                    if cell_length > max_length:

                        max_length = cell_length

                except Exception:

                    pass


            ws.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                45
            )


        # --------------------------------------------------
        # Freeze header row
        # --------------------------------------------------

        ws.freeze_panes = "A2"


        # --------------------------------------------------
        # Save Excel
        # --------------------------------------------------

        wb.save(FILE_NAME)


        print("\n✅ FEEDBACK SAVED!")
        print(
            f"📁 File: {FILE_NAME}"
        )
        print("====================================\n")


        return jsonify({

            "status": "success",

            "message":
                "Feedback saved successfully."

        })


    except Exception as e:

        print("\n❌ ERROR SAVING FEEDBACK")
        print(str(e))


        return jsonify({

            "status": "error",

            "message":
                "Could not save feedback."

        }), 500


# ==========================================================
# RUN SERVER
# ==========================================================

if __name__ == "__main__":

    print("\n====================================")
    print("🤖 EduPulse AI")
    print("====================================")
    print("Server starting...")
    print("Open: http://127.0.0.1:5000")
    print("Excel: data/feedback.xlsx")
    print("====================================\n")

    app.run(
        debug=True,
        host="127.0.0.1",
        port=5000
    ) 